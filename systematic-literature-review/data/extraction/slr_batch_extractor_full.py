import os
import re
import json
import time
from pathlib import Path
from typing import List, Dict, Any

import pandas as pd
from openpyxl import load_workbook
from pypdf import PdfReader
from docx import Document
from openai import OpenAI

# ============================================================
# USER SETTINGS
# ============================================================
PAPERS_DIR = r"/Users/marquemo/Library/CloudStorage/GoogleDrive-mairasamary@gmail.com/.shortcut-targets-by-id/1PnVE0jAHeixRZx6bTu7XkMK724uAntY_/WG10_ITiCSE-2026/Research Efforts/SystematicLiteratureReview/Papers SLR - Phase 2"
TEMPLATE_XLSX = r"Base_table_phase2.xlsx"
SHEET_NAME = "Extraction"
MODEL_NAME = "gpt-4.1"
BATCH_SIZE = 50
MAX_RETRIES = 5
RETRY_BASE_SECONDS = 5
REQUEST_DELAY_SECONDS = 1.0
SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".txt"}
MAX_PAGES_TO_SEND = 40
MAX_CHARS_PER_PAGE = 5000

# Requires environment variable OPENAI_API_KEY
client = OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))

SYSTEM_PROMPT = """
You are an expert research assistant extracting structured data from academic
papers about teamwork in undergraduate computing/engineering education.
Follow the instructions in the user message and return valid JSON only.
"""

# ============================================================
# TEMPLATE / OUTPUT COLUMNS
# ============================================================
def get_headers_from_template(xlsx_path: str, sheet_name: str) -> List[str]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1] if cell.value is not None]
    if not headers:
        raise ValueError(f"No headers found in sheet '{sheet_name}'")
    return headers


def build_output_columns(base_headers: List[str]) -> List[str]:
    output_cols = ["Read Order"]
    for header in base_headers:
        output_cols.append(header)
        if header not in ["Paper Filename", "Drive Link"]:
            output_cols.append(f"{header} - Evidence")
            output_cols.append(f"{header} - Page")
    return output_cols

# ============================================================
# TEXT EXTRACTION
# ============================================================
def normalize_text(text: str) -> str:
    text = text.replace("\x00", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def extract_pdf_pages(file_path: str) -> List[Dict[str, Any]]:
    reader = PdfReader(file_path)
    pages = []
    for i, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception:
            text = ""
        text = normalize_text(text)
        pages.append({"page": i, "text": text[:MAX_CHARS_PER_PAGE]})
    return pages


def extract_docx_pages(file_path: str) -> List[Dict[str, Any]]:
    doc = Document(file_path)
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text and p.text.strip()]
    joined = normalize_text("\n".join(paragraphs))
    chunks = []
    chunk_size = 3500
    current = 0
    pseudo_page = 1
    while current < len(joined):
        chunks.append({"page": pseudo_page, "text": joined[current:current + chunk_size]})
        current += chunk_size
        pseudo_page += 1
    return chunks or [{"page": 1, "text": ""}]


def extract_txt_pages(file_path: str) -> List[Dict[str, Any]]:
    text = Path(file_path).read_text(encoding="utf-8", errors="ignore")
    text = normalize_text(text)
    chunks = []
    chunk_size = 3500
    current = 0
    pseudo_page = 1
    while current < len(text):
        chunks.append({"page": pseudo_page, "text": text[current:current + chunk_size]})
        current += chunk_size
        pseudo_page += 1
    return chunks or [{"page": 1, "text": ""}]


def extract_pages(file_path: str) -> List[Dict[str, Any]]:
    ext = Path(file_path).suffix.lower()
    if ext == ".pdf":
        return extract_pdf_pages(file_path)
    if ext == ".docx":
        return extract_docx_pages(file_path)
    if ext == ".txt":
        return extract_txt_pages(file_path)
    return []


def compress_pages_for_prompt(pages: List[Dict[str, Any]]) -> str:
    pages = pages[:MAX_PAGES_TO_SEND]
    return "\n\n".join([f"[[PAGE {item['page']}]]\n{item['text']}" for item in pages])

# ============================================================
# LLM PROMPTING
# ============================================================
def build_prompt(headers: List[str], filename: str, pages_text: str) -> str:
    question_headers = [h for h in headers if h not in ["Paper Filename", "Drive Link"]]
    return f"""
You are extracting structured evidence from an academic paper for a systematic literature review.

Paper filename: {filename}

You must answer every field below using ONLY the provided paper text.
If unsupported, unclear, or absent, return answer = \"NA\", evidence = \"NA\", and page = \"NA\".
For yes/no fields, answer ONLY with Y, N, or NA.
Do not infer beyond the text.
Use short but precise answers.
Use verbatim evidence snippets copied from the paper when possible.
Use the page number from the page markers [[PAGE X]].
If multiple pages support the same answer, use the strongest single page.

Return valid JSON with exactly this structure:
{{
  "Paper Filename": "{filename}",
  "Drive Link": "",
  "answers": {{
    "FIELD NAME": {{
      "answer": "...",
      "evidence": "...",
      "page": "..."
    }}
  }}
}}

Fields to answer:
{json.dumps(question_headers, ensure_ascii=False, indent=2)}

Paper text:
{pages_text}
"""
#TRYING TO ADD A BETTER MODEL RETRY

from openai import OpenAI, RateLimitError

def call_model_with_retry(prompt):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = client.chat.completions.create(
                model=MODEL_NAME,
                temperature=0,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user", "content": prompt},
                ],
            )
            return response.choices[0].message.content

        except RateLimitError as e:
            msg = str(e)
            if "insufficient_quota" in msg:
                raise RuntimeError(
                    "OpenAI quota exhausted. Add billing/credits or use a different API key."
                ) from e

            if attempt == MAX_RETRIES:
                raise

            sleep_for = 5 * attempt
            print(f"Rate limit hit (attempt {attempt}/{MAX_RETRIES}). Retrying in {sleep_for}s...")
            time.sleep(sleep_for)

        except Exception:
            if attempt == MAX_RETRIES:
                raise
            sleep_for = 5 * attempt
            print(f"Model call failed (attempt {attempt}/{MAX_RETRIES}). Retrying in {sleep_for}s...")
            time.sleep(sleep_for)


## add fail fast
##
##def call_model_with_retry(prompt):
##    # Temporary: no retry, fail fast so we can see the real error
##    response = client.chat.completions.create(
##        model=MODEL_NAME,
##        temperature=0,
##        messages=[
##            {"role": "system", "content": SYSTEM_PROMPT},
##            {"role": "user", "content": prompt},
##        ],
##    )
##    return response.choices[0].message.content
##
##def call_model_with_retry(prompt: str) -> Dict[str, Any]:
##    last_err = None
##    for attempt in range(1, MAX_RETRIES + 1):
##        try:
##            response = client.chat.completions.create(
##                model=MODEL_NAME,
##                temperature=0,
##                response_format={"type": "json_object"},
##                messages=[
##                    {"role": "system", "content": "You extract evidence-grounded structured data from academic papers. Never invent unsupported answers."},
##                    {"role": "user", "content": prompt},
##                ],
##            )
##            return json.loads(response.choices[0].message.content)
##        except Exception as e:
##            last_err = e
##            if attempt == MAX_RETRIES:
##                break
##            sleep_for = RETRY_BASE_SECONDS * attempt
##            print(f"Model call failed (attempt {attempt}/{MAX_RETRIES}). Retrying in {sleep_for}s...")
##            time.sleep(sleep_for)
##    raise last_err

# ============================================================
# ROW BUILDING
# ============================================================
def empty_row(output_columns: List[str], filename: str = "", drive_link: str = "", read_order: Any = "NA") -> Dict[str, Any]:
    row = {col: "NA" for col in output_columns}
    row["Read Order"] = read_order
    row["Paper Filename"] = filename
    row["Drive Link"] = drive_link
    return row


def normalize_yes_no(field_name: str, answer: Any) -> str:
    text = str(answer).strip()
    if "(Y/N)" in field_name:
        upper = text.upper()
        if upper in {"Y", "N", "NA"}:
            return upper
        if upper in {"YES", "TRUE"}:
            return "Y"
        if upper in {"NO", "FALSE"}:
            return "N"
        return "NA"
    return text if text else "NA"


def build_row_from_response(base_headers: List[str], output_columns: List[str], response: Dict[str, Any], filename: str, read_order: int) -> Dict[str, Any]:
    row = empty_row(output_columns, filename=filename, drive_link="", read_order=read_order)
    row["Paper Filename"] = response.get("Paper Filename", filename) or filename
    row["Drive Link"] = response.get("Drive Link", "")
    answers = response.get("answers", {}) if isinstance(response, dict) else {}

    for header in base_headers:
        if header in ["Paper Filename", "Drive Link"]:
            continue
        field_obj = answers.get(header, {}) if isinstance(answers, dict) else {}
        answer = normalize_yes_no(header, field_obj.get("answer", "NA"))
        evidence = str(field_obj.get("evidence", "NA") or "NA").strip()
        page = str(field_obj.get("page", "NA") or "NA").strip()
        row[header] = answer if answer else "NA"
        row[f"{header} - Evidence"] = evidence if evidence else "NA"
        row[f"{header} - Page"] = page if page else "NA"

    return row

# ============================================================
# BATCH SAVING
# ============================================================
def batch_filename(start_idx: int, end_idx: int) -> str:
    return f"papers - {start_idx}-{end_idx}.xlsx"


def save_batch(rows: List[Dict[str, Any]], output_columns: List[str], out_path: str) -> None:
    df = pd.DataFrame(rows)
    df = df.reindex(columns=output_columns)
    df.to_excel(out_path, index=False)

# ============================================================
# MAIN PROCESSING
# ============================================================
def process_all_papers() -> None:
    if not os.environ.get("OPENAI_API_KEY"):
        raise EnvironmentError("Set OPENAI_API_KEY before running this script.")

    papers_dir = Path(PAPERS_DIR)
    if not papers_dir.exists():
        raise FileNotFoundError(f"PAPERS_DIR not found: {PAPERS_DIR}")

    template_path = Path(TEMPLATE_XLSX)
    if not template_path.exists():
        raise FileNotFoundError(f"TEMPLATE_XLSX not found: {TEMPLATE_XLSX}")

    base_headers = get_headers_from_template(str(template_path), SHEET_NAME)
    output_columns = build_output_columns(base_headers)

    paper_files = sorted([
        p for p in papers_dir.iterdir()
        if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS
    ], key=lambda x: x.name.lower())

    if not paper_files:
        raise ValueError("No supported paper files found in PAPERS_DIR.")

    current_batch_rows = []
    current_batch_start = 1

    for global_idx, paper_path in enumerate(paper_files, start=1):
        print(f"[{global_idx}/{len(paper_files)}] Processing: {paper_path.name}")

        try:
            pages = extract_pages(str(paper_path))

            if not pages or not any(p.get("text", "").strip() for p in pages):
                row = empty_row(output_columns, filename=paper_path.name, drive_link="", read_order=global_idx)
                row["IE4: Is the paper available and accessible? (Y/N)"] = "N"
                row["IE4: Is the paper available and accessible? (Y/N) - Evidence"] = "NA"
                row["IE4: Is the paper available and accessible? (Y/N) - Page"] = "NA"
            else:
                prompt = build_prompt(base_headers, paper_path.name, compress_pages_for_prompt(pages))
                #response = call_model_with_retry(prompt)
                #row = build_row_from_response(base_headers, output_columns, response, paper_path.name, global_idx)

                #new
                response_text = call_model_with_retry(prompt).strip()

                if response_text.startswith("```json"):
                    response_text = response_text[7:]
                if response_text.startswith("```"):
                    response_text = response_text[3:]
                if response_text.endswith("```"):
                    response_text = response_text[:-3]

                response_text = response_text.strip()
                response_data = json.loads(response_text)

                row = build_row_from_response(
                    base_headers,
                    output_columns,
                    response_data,
                    paper_path.name,
                    global_idx
                )
                
        except Exception as e:
            row = empty_row(output_columns, filename=paper_path.name, drive_link="", read_order=global_idx)
            row["IE4: Is the paper available and accessible? (Y/N)"] = "NA"
            row["IE4: Is the paper available and accessible? (Y/N) - Evidence"] = f"Processing error: {str(e)[:300]}"
            row["IE4: Is the paper available and accessible? (Y/N) - Page"] = "NA"

        current_batch_rows.append(row)

        batch_is_full = len(current_batch_rows) == BATCH_SIZE
        is_last_paper = global_idx == len(paper_files)

        if batch_is_full or is_last_paper:
            batch_end = global_idx
            out_file = papers_dir / batch_filename(current_batch_start, batch_end)
            save_batch(current_batch_rows, output_columns, str(out_file))
            print(f"Saved batch: {out_file}")
            current_batch_rows = []
            current_batch_start = global_idx + 1

        time.sleep(REQUEST_DELAY_SECONDS)

    print("Done.")


if __name__ == "__main__":
    process_all_papers()
